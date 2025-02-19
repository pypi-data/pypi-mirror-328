import logging
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment


class ExcelReporter:
    def __init__(self) -> None:
        self.wb = Workbook()
        self.logger = logging.getLogger()
    
    def auto_width(self, worksheet: Worksheet):
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    alignment = cell.alignment      # type: Alignment
                    if (alignment.wrapText is True):
                        continue
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

        for col, value in dims.items():
            # worksheet.column_dimensions[col].width = (value + 2) + 2
            worksheet.column_dimensions[col].width = (value + 2)

    def write_title_row(self, sheet: Worksheet, title_row):
        for idx in range(0, len(title_row)):
            cell = sheet.cell(row=1, column=idx + 1)
            cell.value = title_row[idx]

    def write_cell(self, sheet: Worksheet, row: int, column: int, value, format=None) -> Cell:
        cell = sheet.cell(row=row, column=column)         # type: Cell
        cell.value = value
        if (format is not None):
            if ('alignment' in format):
                cell.alignment = format['alignment']
            if ('number_format' in format):
                cell.number_format = format['number_format']
        return cell

    def save(self, name: str):
        self.wb.save(name)
