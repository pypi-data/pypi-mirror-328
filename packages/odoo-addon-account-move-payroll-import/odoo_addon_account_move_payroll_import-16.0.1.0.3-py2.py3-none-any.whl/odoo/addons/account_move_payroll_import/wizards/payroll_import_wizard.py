# -*- coding: utf-8 -*-

import io
import chardet
import datetime
import unicodedata

from odoo import models, fields, _
from odoo.exceptions import UserError
from odoo.tools import (
    pycompat,
    DEFAULT_SERVER_DATE_FORMAT,
    DEFAULT_SERVER_DATETIME_FORMAT,
)
from odoo.addons.base_import.models.base_import import BOM_MAP, ImportValidationError

from ..utils.file_utils import is_valid_extension, decode_file

try:
    import xlrd

    try:
        from xlrd import xlsx
    except ImportError:
        xlsx = None
except ImportError:
    xlrd = xlsx = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class AccountPayrollImportWizard(models.TransientModel):
    _name = "payroll.import.wizard"
    _description = "Account Move Payroll Import Wizard"

    def _default_payroll_import_setup(self):
        payroll_import_setup = self.env["payroll.import.setup"].search([], limit=1)
        if not payroll_import_setup:
            try:
                payroll_import_setup = self.env["payroll.import.setup"].create({})
            except Exception:
                return False

        return payroll_import_setup

    file = fields.Binary(
        string="File", required=True, help=_("Select the payroll file to import.")
    )
    file_name = fields.Char(string="File Name")
    payroll_import_setup_id = fields.Many2one(
        string="Import Configuration",
        comodel_name="payroll.import.setup",
        required=True,
        default=_default_payroll_import_setup,
        help=_("Select the configuration to import the payroll data."),
    )
    account_move_ref = fields.Char(
        string="Account Move Reference",
    )

    def _validate_wizard(self):
        """
        Validate the wizard before importing the file.
        - Check if the journal is set in the import configuration:
            The journal is required in the UI, but not in the model definition,
            so we need to check it is set.
        - Check if the file extension is supported.
        - Check if the 'xlrd' library is installed to read XLS or XLSX files.
        :return: the file extension.
        """
        if not self.payroll_import_setup_id.journal_id:
            raise UserError(_("Please select a journal in the import configuration."))

        for custom_concept in self.payroll_import_setup_id.custom_concepts_ids:
            if not custom_concept.account_id:
                raise UserError(
                    _(
                        "Please select an account for the custom concept: %s."
                        % custom_concept.name
                    )
                )

        extension = is_valid_extension(self.file_name)
        if not extension:
            raise UserError(
                _("Unsupported file format. Import only supports CSV, XLS and XLSX.")
            )

        if extension in ["xls", "xlsx"] and not xlrd and not load_workbook:
            raise UserError(
                _(
                    "Please install the 'xlrd' or 'openpyxl' library to import XLS files."
                )  # noqa
            )

        return extension

    def _read_xls(self, options):
        """
        Read the XLS or XLSX file content.
        :param file: the file content to read.

        - See also: odoo/addons/base_import/models/base_import.py
        """
        book = xlrd.open_workbook(file_contents=decode_file(self.file) or b"")
        sheets = options["sheets"] = book.sheet_names()
        sheet_name = options["sheet"] = options.get("sheet") or sheets[0]

        sheet = book.sheet_by_name(sheet_name)
        rows = []
        # emulate Sheet.get_rows for pre-0.9.4
        for rowx, row in enumerate(map(sheet.row, range(sheet.nrows)), 1):
            values = []
            for colx, cell in enumerate(row, 1):
                if cell.ctype is xlrd.XL_CELL_NUMBER:
                    is_float = cell.value % 1 != 0.0
                    values.append(
                        str(cell.value) if is_float else str(int(cell.value))
                    )
                elif cell.ctype is xlrd.XL_CELL_DATE:
                    is_datetime = cell.value % 1 != 0.0
                    # emulate xldate_as_datetime for pre-0.9.3
                    dt = datetime.datetime(
                        *xlrd.xldate.xldate_as_tuple(cell.value, book.datemode)
                    )
                    values.append(
                        dt.strftime(DEFAULT_SERVER_DATETIME_FORMAT)
                        if is_datetime
                        else dt.strftime(DEFAULT_SERVER_DATE_FORMAT)
                    )
                elif cell.ctype is xlrd.XL_CELL_BOOLEAN:
                    values.append("True" if cell.value else "False")
                elif cell.ctype is xlrd.XL_CELL_ERROR:
                    raise ValueError(
                        _(
                            "Invalid cell value at row %(row)s, column %(col)s: %(cell_value)s"
                        )
                        % {
                            "row": rowx,
                            "col": colx,
                            "cell_value": xlrd.error_text_from_code.get(
                                cell.value, _("unknown error code %s", cell.value)
                            ),
                        }
                    )
                else:
                    values.append(cell.value)

            rows.append(values)

        return rows

    # use the same method for xlsx and xls files
    def _read_xlsx(self, options):
        if xlsx:
            return self._read_xls(options)

        import openpyxl.cell.cell as types
        import openpyxl.styles.numbers as styles  # noqa: PLC0415

        book = load_workbook(io.BytesIO(decode_file(self.file) or b""), data_only=True)
        sheets = options["sheets"] = book.sheetnames
        sheet_name = options["sheet"] = options.get("sheet") or sheets[0]
        sheet = book[sheet_name]
        rows = []
        for rowx, row in enumerate(sheet.rows, 1):
            values = []
            for colx, cell in enumerate(row, 1):
                if cell.data_type is types.TYPE_ERROR:
                    raise ValueError(
                        _(
                            "Invalid cell value at row %(row)s, column %(col)s: %(cell_value)s",
                            row=rowx,
                            col=colx,
                            cell_value=cell.value,
                        )
                    )

                if cell.value is None:
                    values.append("")
                elif isinstance(cell.value, float):
                    if cell.value % 1 == 0:
                        values.append(str(int(cell.value)))
                    else:
                        values.append(str(cell.value))
                elif cell.is_date:
                    d_fmt = styles.is_datetime(cell.number_format)
                    if d_fmt == "datetime":
                        values.append(
                            cell.value.strftime(DEFAULT_SERVER_DATETIME_FORMAT)
                        )
                    elif d_fmt == "date":
                        values.append(cell.value.strftime(DEFAULT_SERVER_DATE_FORMAT))
                    else:
                        raise ValueError(
                            _(
                                "Invalid cell format at row %(row)s, column %(col)s: %(cell_value)s, with format: %(cell_format)s, as (%(format_type)s) formats are not supported.",  # noqa
                                row=rowx,
                                col=colx,
                                cell_value=cell.value,
                                cell_format=cell.number_format,
                                format_type=d_fmt,
                            )
                        )
                else:
                    values.append(str(cell.value))

            rows.append(values)
        return rows

    def _read_csv(self, options):
        """Returns file length and a CSV-parsed list of all non-empty lines in the file.

        :raises csv.Error: if an error is detected during CSV parsing
        """
        csv_data = decode_file(self.file) or b""
        if not csv_data:
            return []

        encoding = options.get("encoding")
        if not encoding:
            encoding = options["encoding"] = chardet.detect(csv_data)[
                "encoding"
            ].lower()
            # some versions of chardet (e.g. 2.3.0 but not 3.x) will return
            # utf-(16|32)(le|be), which for python means "ignore / don't strip
            # BOM". We don't want that, so rectify the encoding to non-marked
            # IFF the guessed encoding is LE/BE and csv_data starts with a BOM
            bom = BOM_MAP.get(encoding)
            if bom and csv_data.startswith(bom):
                encoding = options["encoding"] = encoding[:-2]

        if encoding != "utf-8":
            csv_data = csv_data.decode(encoding).encode("utf-8")

        separator = options.get("separator")
        if not separator:
            # default for unspecified separator so user gets a message about
            # having to specify it
            separator = ","
            for candidate in (
                ",",
                ";",
                "\t",
                " ",
                "|",
                unicodedata.lookup("unit separator"),
            ):
                # pass through the CSV and check if all rows are the same
                # length & at least 2-wide assume it's the correct one
                it = pycompat.csv_reader(
                    io.BytesIO(csv_data),
                    quotechar=options["quoting"],
                    delimiter=candidate,
                )
                w = None
                for row in it:
                    width = len(row)
                    if w is None:
                        w = width
                    if width == 1 or width != w:
                        break  # next candidate
                else:  # nobreak
                    separator = options["separator"] = candidate
                    break

        if not len(options["quoting"]) == 1:
            raise ImportValidationError(
                _(
                    "Error while importing records: Text Delimiter should be a single character."  # noqa
                )
            )

        csv_iterator = pycompat.csv_reader(
            io.BytesIO(csv_data), quotechar=options["quoting"], delimiter=separator
        )

        return [row for row in csv_iterator]

    def _read_file(self, extension):
        """
        Reading based on the file extension.
        """
        import_setup = self.payroll_import_setup_id
        options = import_setup.get_file_options()

        rows, cumulative_tc1rlc_ss = [], 0.0
        skip_lines = options.get("header_lines", 1) - 1
        for idx, row in enumerate(getattr(self, f"_read_{extension}")(options)):
            if idx == 0:
                import_setup.validate_column_numbers(len(row))
            if idx == import_setup.header_ref_line - 1:
                self.account_move_ref = row[0].strip() or import_setup.name

            if idx < skip_lines:
                continue
            row = [str(col).strip() for col in row]

            # exclude empty rows or rows without employee id (summary rows)
            if any(row) and row[import_setup.column_employee_id - 1]:
                rows.append(list(row))

                # compute cumulative total tc1rlc minus ss_employee and ss_columns
                cumulative_tc1rlc_ss = import_setup.compute_tc1rlc_ss_cumulative(
                    row, cumulative_tc1rlc_ss
                )

        if cumulative_tc1rlc_ss > 0.0:
            raise UserError(
                _(
                    "TC1 total should be equal to the sum of the SS Employee"
                    " and SS Company totals. Please check the file."
                )
            )

        return rows

    def action_import(self):
        extension = self._validate_wizard()

        file_content = self._read_file(extension)
        if not file_content:
            raise UserError(_("The file is empty."))

        move = self.payroll_import_setup_id.process_data(file_content)
        move.ref = self.account_move_ref

        return {
            "type": "ir.actions.client",
            "tag": "reload",
        }
