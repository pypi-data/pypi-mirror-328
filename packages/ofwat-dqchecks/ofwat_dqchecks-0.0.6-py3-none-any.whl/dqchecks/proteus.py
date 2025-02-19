import datetime
import re
import pandas as pd
from openpyxl.styles import PatternFill
import numpy as np

green_fill = PatternFill(start_color='72C931', end_color='72C931', fill_type='solid')
red_fill = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')

##Import Company excel file
def import_data(file_location, excel_error_log_name):
    firstsheet = excel_error_log_name['Intro']
    print('File Location Name : ', file_location)
    try:
        xl1 = pd.ExcelFile(file_location)  # Define the excel file
        worksheets = xl1.sheet_names  # Get the list of worksheets in the file
        found_dictionary, found_f_outputs = (False, False)
        dict_of_sheets = {}
        for sheet in worksheets:  # iterate through the sheets in the file
            if sheet.startswith("Dict_"):  # ...if the worksheet starts with the text "Dict_" then...
                found_dictionary = True
                dict_of_sheets[sheet] = (pd.read_excel(xl1, sheet_name=sheet, skiprows=[0, 2]), "Dict_")
            if sheet.startswith("fOut_"):  # ...if the workseet starts with the text "fOut" then...
                found_f_outputs = True
                dict_of_sheets[sheet] = (pd.read_excel(xl1, sheet_name=sheet, skiprows=[0, 2]), "fOut_")
        if not found_f_outputs:
            firstsheet['B6'].value = f"{firstsheet['B6'].value} Error message: A worksheet starting with 'fOut_' (an F_Outputs sheet) was not found in worksheets, Proteus has failed to run"
            firstsheet['B6'].fill = red_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError}"
        firstsheet['B6'].fill = red_fill
        pass
    return dict_of_sheets

##Import original excel file
def import_original_data(original_file_location, excel_error_log_name):
    firstsheet = excel_error_log_name['Intro']
    print('original_file_location : ', original_file_location)
    try:
        xl1 = pd.ExcelFile(original_file_location)  # Define the excel file
        worksheets = xl1.sheet_names  # Get the list of worksheets in the file
        original_dict_of_sheets = {}
        for sheet in worksheets:  # iterate through the sheets in the file
            if sheet.startswith("Dict_"):  # ...if the workseet starts with the text "Dict_" then...
                original_dict_of_sheets[sheet] = (pd.read_excel(xl1, sheet_name=sheet, skiprows=[0, 2]), "Dict_")
            if sheet.startswith("fOut_"):  # ...if the workseet starts with the text "fOut" then...
                excel_error_log_name.create_sheet(title=sheet[5:] + ' Output')  # creating the sheets which will house the tables of outputs
                original_dict_of_sheets[sheet] = (pd.read_excel(xl1, sheet_name=sheet, skiprows=[0, 2]), "fOut_")
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError}"
        firstsheet['B6'].fill = red_fill
        pass
    return original_dict_of_sheets

def timestamp_error_log(excel_error_log_name):
    firstsheet = excel_error_log_name['Intro']
    try:
        # storing current date and time
        current_date_time = datetime.datetime.now().strftime("%d.%b %Y %H:%M:%S")
        firstsheet['B4'].value = f"This file was created at: {current_date_time}"
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError}"
        firstsheet['B6'].fill = red_fill
        pass

def fOut_headers_consistency(sheet_name, df, excel_error_log_name):
    ## ---------------------------------------------------------------------
    ## *Headers* in spesific order and value. Headers are case insensitive
    ## ----------------------------------------------------------------------
    firstsheet = excel_error_log_name['Intro']
    try:
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output'] #defining which sheet the output needs to go into
        outputsheet['B3'].value = 'Data Validation Rule 1'
        outputsheet['C3'].value = ""

        dictionary_errors_column_headers = {
            'Acronym': (df.columns[0],
                        "Error message: 'Acronym' column name is not correct, please note name is case sensitive"),
            'Reference': (df.columns[1],
                          "Error message: 'Reference' column name is not correct, please note name is case sensitive"),
            'Item description': (df.columns[2],
                                 "Error message: 'Item description' column name is not correct, please note name is case sensitive"),
            'Unit': (
            df.columns[3], "Error message: 'Unit' column name is not correct, please note name is case sensitive"),
            'Model': (
            df.columns[4], "Error message: 'Model' column name is not correct, please note name is case sensitive"),
        }

        error_counter = 0
        for key, value in dictionary_errors_column_headers.items():
            if key == value[0]:
                pass
            else:
                error_counter += 1
                outputsheet['C3'].value = f"{outputsheet['C3'].value} {value[1]}"
                outputsheet['C3'].fill = red_fill
        if error_counter == 0:
            outputsheet['C3'].value = "Success, No Errors detected!"
            outputsheet['C3'].fill = green_fill
            return True
        outputsheet['C3'].value = f"{outputsheet['C3'].value} Please correct the header's names for this Excel Sheet: {sheet_name}, before data validation proceeds!"
        outputsheet['C3'].fill = red_fill
        return False
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError}"
        firstsheet['B6'].fill = red_fill
        return False

def regex_boncode(sheet_name,df,excel_error_log_name):
##-------------------------------------------------------------------------------------------------------
##Boncode find pattern using *Regex*. Reference (BON code) should consist of the following in sequence:
##An upper case letter. Zero or more upper case letters or underscores.
## At least one digit. Zero or more number upper case letters, underscores, or digits.
##No lower case letters. Boncodes must not contain "-" character.
##Note "str.fullmatch" doesn't work due to older python vesrion installed
##-------------------------------------------------------------------------------------------------------
    firstsheet = excel_error_log_name['Intro']
    try:
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output']  # defining which sheet the output needs to go into
        outputsheet['B4'].value = 'Data Validation Rule 2'
        outputsheet['C4'].value = ""
        index=3 #index for showing cell position
        regex = r'^([A-Z][A-Z_]*[0-9]+[A-Z0-9_]*)$'
        regex_pattern = re.compile(regex)
        error_counter=0
        for item in df['Reference']:
            index+=1
            if pd.isna(item):
                error_counter+=1
                outputsheet['C4'].value = f"{outputsheet['C4'].value} Error in Row {index}: 'Reference': {item}, 'Reference' is a mandatory field and cannot have empty values \n"
                outputsheet['C4'].fill = red_fill
                continue  # Skip further processing for this row
            if not isinstance(item, str):
                error_counter += 1
                outputsheet['C4'].value = f"{outputsheet['C4'].value} Error in Row {index}: 'Reference': {item}, 'Reference' should be a string \n"
                outputsheet['C4'].fill = red_fill
                continue 
            if not regex_pattern.match(item):
                error_counter+=1
                outputsheet['C4'].value = f"{outputsheet['C4'].value} Error in Row {index}: 'Reference': {item} doesn't match the regular expression \n"
                outputsheet['C4'].fill = red_fill
        if error_counter==0:
            outputsheet['C4'].value = "Success, No Errors detected!"
            outputsheet['C4'].fill = green_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError} \n"
        firstsheet['B6'].fill = red_fill
        pass

def check_suffix(sheet_name, df, excel_error_log_name):
    ## ---------------------------------------
    ## Checking *suffix _PR24* on Boncodes
    ## ---------------------------------------
    firstsheet = excel_error_log_name['Intro']
    try:
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output']  # defining which sheet the output needs to go into
        outputsheet['B5'].value = 'Data Validation Rule 3'
        outputsheet['C5'].value = ""
        index = 3  # index for showing row position
        suffix = "_PR24"
        error_counter = 0

        for item in df['Reference']:
            index += 1
            if pd.isna(item):
                error_counter += 1
                outputsheet['C5'].value = f"{outputsheet['C5'].value} Error in Row {index}: Reference is missing (NaN)\n"
                outputsheet['C5'].fill = red_fill
            elif isinstance(item, str):
                if not item.endswith(suffix):
                    error_counter += 1
                    outputsheet['C5'].value = f"{outputsheet['C5'].value} Error in Row {index}: Reference: {item} does not have the suffix _PR24\n"
                    outputsheet['C5'].fill = red_fill
            else:
                error_counter += 1
                outputsheet['C5'].value = f"{outputsheet['C5'].value} Error in Row {index}: Reference: {item} is not a valid string\n"
                outputsheet['C5'].fill = red_fill
        if error_counter == 0:
            outputsheet['C5'].value ="Success, No Errors detected!"
            outputsheet['C5'].fill = green_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError} \n"
        firstsheet['B6'].fill = red_fill
        pass

def check_data_type(sheet_name, df, excel_error_log_name):
    ## ------------------------------------------------------------------------------------------
    ##6. *Data Type Checking*. This verifies that the entered data has the appropriate data type.
    ## ------------------------------------------------------------------------------------------
    # Unit: Unit must be less than 21 characters.
    # Description: Description must be less than 230 characters.
    regex_columns = r'^([0-9]{4}-[0-9]{2})$'
    regex_pattern = re.compile(regex_columns)
    firstsheet = excel_error_log_name['Intro']
    try:
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output']  # defining which sheet the output needs to go into
        outputsheet['B6'].value = 'Data Validation Rule 4'
        outputsheet['C6'].value = ""
        index = 3  # index for showing row position
        error_counter = 0
        unit_length = 21
        description_length = 230

        for item1, item2, item3 in zip(df['Unit'], df['Reference'],df['Item description']):
            index += 1
            if isinstance(item1, int):
                error_counter += 1
                outputsheet['C6'].value = f"{outputsheet['C6'].value} Error in Row {index}: Reference: {item2}: Unit must not be a number \n"
                outputsheet['C6'].fill = red_fill
            elif pd.notna(item1) and isinstance(item1, str):
                if len(item1) > unit_length:
                    error_counter += 1
                    outputsheet['C6'].value = f"{outputsheet['C6'].value} Error in Row {index}: Reference: {item2}: Unit must be <={unit_length} characters \n"
                    outputsheet['C6'].fill = red_fill
            elif pd.notna(item3) and isinstance(item3, str):
                if len(item3) > description_length:
                    error_counter += 1
                    outputsheet['C6'].value = f"{outputsheet['C6'].value} Error in Row {index}: Reference: {item2}: Description must be <={description_length} characters \n"
                    outputsheet['C6'].fill = red_fill
        index = 3
        for i, row in df.iterrows():
            index += 1
            for key in row.keys():
                if bool(regex_pattern.match(key)) or key.lower() == 'constant':
                    if len(str(row[key])) > description_length:
                        error_counter +=1
                        outputsheet['C6'].value = f"{outputsheet['C6'].value}Error in Row: {index}, Reference: {df.loc[i, 'Reference']},Column: {key}, Value: {row[key]}- this value is over 230 characters.\n\n"
                        outputsheet['C6'].fill = red_fill
        if error_counter == 0:
            outputsheet['C6'].value = "Success, No Errors detected!"
            outputsheet['C6'].fill = green_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError} \n"
        firstsheet['B6'].fill = red_fill
        pass

def check_ref(sheet_name, df, excel_error_log_name):
    # REF! error
    firstsheet = excel_error_log_name['Intro']
    regex_columns = r'^([0-9]{4}-[0-9]{2})$'
    regex_pattern = re.compile(regex_columns)
    
    try:
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output']  # defining which sheet the output needs to go into
        outputsheet['B7'].value = 'Data Validation Rule 5'
        outputsheet['C7'].value = ""
        index = 3
        error_counter = 0

        for i, row in df.iterrows():
            index += 1
            if row['Model'] and pd.isna(row['Reference']) and pd.isna(row['Unit']):
                error_counter += 1
                outputsheet['C7'].value = f"{outputsheet['C7'].value} Error in Row: {index}," \
                    "This likely contains a reference error, please check \n"
                outputsheet['C7'].fill = red_fill
        if error_counter == 0:
            outputsheet['C7'].value = "Success, No Errors detected!"
            outputsheet['C7'].fill = green_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {str(ValueError)} \n"
        firstsheet['B6'].fill = red_fill
        pass

def boncode_Uniqueness(sheet_name,df,excel_error_log_name):
#-----------------------------------------------------------------------------------
#Boncodes Uniqueness*: same boncode has been used only once in the dictionary file 
#Identify any duplicated Boncodes in across the dictionary
#-----------------------------------------------------------------------------------
#add items one by one to a list and while adding check if it is duplicated or not
    firstsheet = excel_error_log_name['Intro']
    try: 
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output']  # defining which sheet the output needs to go into
        outputsheet['B10'].value = 'Data Validation Rule 8'
        index=3 #index for showing row position
        boncode_duplicates = []
        error_counter=0
        
        for item in df['Reference']:  
            index += 1
            if item in boncode_duplicates:
                error_counter += 1
                outputsheet['C10'].value = f"Error in row {index}: Reference: {item} is not unique"
                outputsheet['C10'].fill = red_fill
            else: 
                boncode_duplicates.append(item)
        if error_counter == 0:
            outputsheet['C10'].value = "Success, No Errors detected!"
            outputsheet['C10'].fill = green_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value}{ValueError} \n"
        firstsheet['B6'].fill = red_fill
        pass

def user_text_input(sheet_name, df, excel_error_log_name):
    # ---------------------------------------------------------------------
    ##Whether the values (e.g. anything under the years headers) corresponds to what is in the units columns
    # If the Unit value is not ‘Text’ the years columns values should be number
    #sheet RR 414-419, 489-491 error because unit is weird
    # ---------------------------------------------------------------------
    regex_columns = r'^([0-9]{4}-[0-9]{2})$'
    regex_pattern = re.compile(regex_columns)

    firstsheet = excel_error_log_name['Intro']
    try:
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output']  # defining which sheet the output needs to go into
        outputsheet['B8'].value = 'Data Validation Rule 6 (will not stop data loading to fountain)'
        outputsheet['C8'].value = ""
        index = 3
        error_counter = 0

        for i, row in df.iterrows():
            index += 1
            for key in row.keys():
                if bool(regex_pattern.match(key)) or key.lower() == 'constant':
                    if row[key] == "##BLANK":
                        continue
                    elif row[key] == True:
                        continue
                    elif row[key] == False:
                        continue
                    elif row[key] == "#REF!": 
                        continue
                    elif (row["Unit"] == "Time"):
                        continue
                    elif row[key] == "#DIV/0!":
                        continue
                    elif pd.isna(row[key]):
                        continue
                    elif row[key] == '\xa0':
                        continue
                    elif row[key] == int(0) or row[key] == float(0) or row[key] == str(0):
                        continue
                    elif pd.isna(row[key]) or pd.isna(row["Unit"]):
                        continue
                    # if unit is text and value in every column is not string
                    #elif (row["Unit"].lower() in ["text"]) and type(row[key]) != str:
                    elif isinstance(row["Unit"], str) and row["Unit"].lower() == "text" and not isinstance(row[key], str):
                        error_counter += 1
                        outputsheet['C8'].value = f"{outputsheet['C8'].value} Error in Row: {index}, Reference: {df.loc[i, 'Reference']}, Column: {key}, Value: {row[key]}, " \
                            "This is a text field, please check that your input is in text format \n"
                        outputsheet['C8'].fill = red_fill
        if error_counter == 0:
            outputsheet['C8'].value = "Success, No Errors detected!"
            outputsheet['C8'].fill = green_fill 
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError} \n"
        firstsheet['B6'].fill = red_fill
        pass

def boncode_Consistency(sheet_name,dict_of_sheets,original_dict_of_sheets,excel_error_log_name):

#-----------------------------------------------------------------------------------
#Boncode Consistency*, prevent accidental change/delete of Boncodes
#-----------------------------------------------------------------------------------
    firstsheet = excel_error_log_name['Intro']
    try:
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output']  # defining which sheet the output needs to go into
        outputsheet['B9'].value = 'Data Validation Rule 7'
        outputsheet['C9'].value = ""
        amended_boncodes=dict_of_sheets['Reference']
        original_boncodes=original_dict_of_sheets['Reference']
        result_string = ''
        original_only_boncodes = list(set(original_boncodes) - set(amended_boncodes))  # The items in the original version, but not in ammended version
        for item in original_only_boncodes:
            if item !=np.nan and item != 'nan':
                result_string += str(item) + '\n'
        if len(original_only_boncodes) > 0:
            outputsheet['C9'].value = f"{outputsheet['C9'].value} Reference (Boncodes) in worksheet in the original version but not in the amended version: {result_string}"
            outputsheet['C9'].fill = red_fill
        else:
            outputsheet['C9'].value = "Success, No Errors detected!"
            outputsheet['C9'].fill = green_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError} \n"
        firstsheet['B6'].fill = red_fill
        pass

def user_numerical_input(sheet_name, df, excel_error_log_name):
    # ---------------------------------------------------------------------
    ##Whether the values (e.g. anything under the years headers) corresponds to what is in the units columns
    # If the Unit value is not ‘Text’ the years columns values should be number
    #sheet RR 414-419, 489-491 error because unit is weird
    # ---------------------------------------------------------------------
    unit_list = [ '£','000','days','hours','kg','km','kw','l/', 'litres','m2','m3','mg','minutes','ml','months','mtrs','num','tonnes','year','ttds']

    regex_columns = r'^([0-9]{4}-[0-9]{2})$'
    regex_pattern = re.compile(regex_columns)

    firstsheet = excel_error_log_name['Intro']
    try:
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output']  # defining which sheet the output needs to go into
        outputsheet['B11'].value = 'Data Validation Rule 9'
        outputsheet['C11'].value = ""

        index = 3  # index for showing row position
        error_counter = 0

        for i, row in df.iterrows():
            index += 1
            for key in row.keys():
                if bool(regex_pattern.match(key)) or key.lower() == 'constant':
                    if row[key] == True:
                        continue
                    elif row[key] == False: 
                        continue
                    elif row[key] == "##BLANK":
                        continue
                    elif row[key] == "#DIV/0!":
                        continue
                    elif row[key] == "#REF!": 
                        continue
                    elif pd.isna(row[key]):
                        continue
                    elif row[key] == '\xa0':
                        continue
                    elif row[key] == int(0) or row[key] == float(0) or row[key] == str(0):
                        continue
                    elif pd.isna(row["Unit"]):
                        continue
                    elif (row["Unit"] == "Time"):
                        continue
                    # if unit is numeric, look at the list and value in every column is not int or float
                    else: 
                        for substring in unit_list:
                            if substring in str(row["Unit"]).lower():
                                if isinstance(row[key], str) and row[key].replace('.', '').isdigit():
                                    continue
                                elif type(row[key]) != int and type(row[key]) != float:
                                    #print(f"type: {type(row[key])}, value: {row[key]}, bon: {row['Reference']}, year: {key}  ")
                                    error_counter += 1
                                    outputsheet['C11'].value = f"{outputsheet['C11'].value} Error in Row: {index}, Reference: {df.loc[i, 'Reference']}, Column: {key}, Value: {row[key]}, " \
                                        "This is a numeric field, please check that your input is in numerical format. \n"
                                    outputsheet['C11'].fill = red_fill
        if error_counter == 0:
            outputsheet['C11'].value = "Success, No Errors detected!"
            outputsheet['C11'].fill = green_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError} \n"
        firstsheet['B6'].fill = red_fill
        pass


def percentage_range(sheet_name, df, excel_error_log_name):
    #checks that entries where the unit is a percentage fall between 0 and 100
    regex_columns = r'^([0-9]{4}-[0-9]{2})$'
    regex_pattern = re.compile(regex_columns)

    firstsheet = excel_error_log_name['Intro']
    try:
        outputsheet = excel_error_log_name[sheet_name[5:] + ' Output']  # defining which sheet the output needs to go into
        outputsheet['B12'].value = 'Data Validation Rule 10 (will not stop data loading to fountain)'
        outputsheet['C12'].value = ""

        index = 3  # index for showing row position
        error_counter = 0

        for i, row in df.iterrows():
            index += 1
            for key in row.keys():
                if bool(regex_pattern.match(key)) or key.lower() == 'constant':
                    if (row["Unit"] == "%"):
                        if type(row[key]) == int or type(row[key]) == float:
                            if row[key]<-2 or row[key]>2:
                                error_counter += 1
                                outputsheet['C12'].value = f"{outputsheet['C12'].value} Error in Row: {index}, Reference: {df.loc[i, 'Reference']}, Column: {key}, Value: {row[key]}, " \
                                    "This is a percentage outside the expected range, with absolute value less than 2, please check that this value is correct. \n"
                                outputsheet['C12'].fill = red_fill        
        if error_counter == 0:
            outputsheet['C12'].value = "Success, No Errors detected!"
            outputsheet['C12'].fill = green_fill
    except ValueError:
        firstsheet['B6'].value = f"{firstsheet['B6'].value} {ValueError} \n"
        firstsheet['B6'].fill = red_fill
        pass