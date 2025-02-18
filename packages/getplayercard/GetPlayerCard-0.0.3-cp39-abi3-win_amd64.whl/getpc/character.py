import json
import random
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from typing import Dict, Any, List

class CharacterGenerator:
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.validate_config()

    def validate_config(self):
        """验证配置文件有效性"""
        required_sections = ['attributes', 'races', 'classes', 'backgrounds']
        for section in required_sections:
            if section not in self.config:
                raise ValueError(f"Missing required config section: {section}")

    def roll_value(self, roll_rule: str) -> int:
        """执行掷骰规则"""
        if roll_rule == "4d6d1":
            rolls = sorted([random.randint(1,6) for _ in range(4)], reverse=True)[:3]
            return sum(rolls)
        elif roll_rule.startswith("d"):
            max_val = int(roll_rule[1:])
            return random.randint(1, max_val)
        # 可以扩展更多掷骰规则
        return 0

    def create_character(self, custom_data: Dict[str, Any] = None) -> Dict[str, Any]:
        """创建新角色"""
        char_data = {
            "attributes": {},
            "skills": [],
            "equipment": [],
            **custom_data or {}
        }

        # 生成基础属性
        for attr, rule in self.config['attributes'].items():
            char_data['attributes'][attr] = self.roll_value(rule['roll_rule'])

        # 应用种族加成
        selected_race = char_data.get('race') or random.choice(list(self.config['races'].keys()))
        race_bonus = self.config['races'][selected_race].get('attribute_bonus', {})
        for attr, bonus in race_bonus.items():
            char_data['attributes'][attr] += bonus

        # 添加职业装备
        selected_class = char_data.get('class') or random.choice(list(self.config['classes'].keys()))
        char_data['equipment'] = self.config['classes'][selected_class].get('equipment', [])

        # 添加背景技能
        selected_bg = char_data.get('background') or random.choice(list(self.config['backgrounds'].keys()))
        char_data['skills'] = self.config['backgrounds'][selected_bg].get('skills', [])

        return char_data

class ExcelExporter:
    def __init__(self, template_config: Dict[str, Any]):
        self.template_config = template_config

    def apply_style(self, sheet):
        """应用单元格样式"""
        if 'styles' in self.template_config:
            for cell_ref, style in self.template_config['styles'].items():
                cell = sheet[cell_ref]
                if 'font' in style:
                    cell.font = Font(**style['font'])
                if 'alignment' in style:
                    cell.alignment = Alignment(**style['alignment'])

    def export(self, characters: List[Dict[str, Any]], filename: str):
        """导出到Excel文件"""
        wb = openpyxl.Workbook()
        sheet = wb.active

        # 写入表头
        for col, field in enumerate(self.template_config['fields'], 1):
            sheet.cell(row=1, column=col, value=field['display_name'])
            if 'width' in field:
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = field['width']

        # 写入数据
        for row, char in enumerate(characters, 2):
            for col, field in enumerate(self.template_config['fields'], 1):
                value = self._get_nested_value(char, field['field_path'])
                sheet.cell(row=row, column=col, value=value)

        self.apply_style(sheet)
        wb.save(filename)

    def _get_nested_value(self, data: Dict, path: str) -> Any:
        """获取嵌套字典值"""
        keys = path.split('.')
        value = data
        for key in keys:
            value = value.get(key, '')
            if not isinstance(value, dict):
                break
        return value if value is not None else ''

# 配置文件示例 (可保存为JSON/YAML)
CONFIG_EXAMPLE = {
    "attributes": {
        "力量": {
            "roll_rule": "4d6d1",
            "display": "STR"
        },
        "魔力": {
            "roll_rule": "2d6+6",
            "display": "MAG"
        }
    },
    "races": {
        "精灵": {
            "attribute_bonus": {"敏捷": 2, "智力": 1},
            "traits": ["黑暗视觉"]
        }
    },
    "classes": {
        "魔剑士": {
            "equipment": ["魔法剑", "轻甲"]
        }
    },
    "backgrounds": {
        "皇室成员": {
            "skills": ["礼仪", "政治"]
        }
    }
}

TEMPLATE_CONFIG_EXAMPLE = {
    "fields": [
        {
            "field_path": "name",
            "display_name": "角色名",
            "width": 20
        },
        {
            "field_path": "attributes.力量",
            "display_name": "力量(STR)",
            "width": 15
        }
    ],
    "styles": {
        "A1:Z1": {
            "font": {"bold": True, "color": "FFFFFF"},
            "alignment": {"horizontal": "center"}
        }
    }
}

# 使用示例
if __name__ == "__main__":
    # 加载配置文件
    with open('character_config.json') as f:
        config = json.load(f)
    
    # 初始化生成器
    generator = CharacterGenerator(config)
    
    # 创建角色
    character = generator.create_character({
        "name": "艾丽卡",
        "race": "精灵",
        "class": "魔剑士"
    })
    
    # 导出到Excel
    exporter = ExcelExporter(template_config=TEMPLATE_CONFIG_EXAMPLE)
    exporter.export([character], "custom_character.xlsx")