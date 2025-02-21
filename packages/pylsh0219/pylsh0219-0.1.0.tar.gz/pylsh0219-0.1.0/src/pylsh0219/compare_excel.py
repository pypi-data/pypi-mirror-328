"""步骤三：比较两个表格中的礼包内容和累充需求，并将对比结果保存到新的 Excel 文件中。"""
import re
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 加载过滤规则
with open('filter_rules.json', 'r', encoding='utf-8') as f:
    filter_rules = json.load(f)

def parse_gift_content(content):
    """
    解析礼包内容，将其拆分为单词和数字的组合，并放入单独的单元格中。
    """
    # 替换中文符号为英文符号
    for old, new in filter_rules["replacements"].items():
        content = content.replace(old, new)

    # 去掉所有的 * 和 x
    for char in filter_rules["remove_chars"]:
        content = content.replace(char, '')

    # 去掉括号内的逗号
    content = re.sub(filter_rules["regex_patterns"]["parentheses"], lambda match: match.group(0).replace(',', ''), content)

    # 将内容按逗号、顿号、空格等分隔符拆分为列表
    items = [item.strip() for item in content.split(',')]
    # 过滤空字符串
    items = [item for item in items if item]
    return items

def parse_tip_content(tip_content):
    """
    解析 tip 内容，提取是否重复和兑换金额信息。
    """
    # 是否重复
    repeat_info = "未知"
    for key, keywords in filter_rules["repeat_keywords"].items():
        for keyword in keywords:
            if keyword in tip_content:
                repeat_info = key
                break

    # 兑换金额
    amount_match = re.search(filter_rules["regex_patterns"]["amount"], tip_content)
    if amount_match:
        amount = int(amount_match.group(1))  # 转换为整数
    else:
        amount = 0 if "免费" in tip_content else -1  # 免费为 0，未知为 -1

    return repeat_info, amount

def is_partial_match(ocr_item, target_item):
    """
    检查 OCR 项是否是目标项的部分匹配（前缀或后缀），忽略标点符号和空格。
    """
    # 去掉标点符号和空格
    ocr_item_clean = re.sub(r'[^\w]', '', ocr_item).lower()
    target_item_clean = re.sub(r'[^\w]', '', target_item).lower()
    return ocr_item_clean in target_item_clean or target_item_clean in ocr_item_clean

def compare_gift_content(ocr_items, target_items):
    """
    比较目标内容是否是 OCR 内容的子集，允许部分匹配和拼接 OCR 内容项，忽略文本格式差异。
    """
    # 检查目标内容的每一项是否都存在于 OCR 内容中，或者可以通过部分匹配或拼接 OCR 内容项得到
    diff_items = []
    for target_item in target_items:
        # 去掉标点符号和空格
        target_item_clean = re.sub(r'[^\w]', '', target_item).lower()
        found = False
        for ocr_item in ocr_items:
            ocr_item_clean = re.sub(r'[^\w]', '', ocr_item).lower()
            if is_partial_match(ocr_item_clean, target_item_clean):
                found = True
                break
        # 如果部分匹配未找到，尝试拼接 OCR 内容项
        if not found:
            for i in range(len(ocr_items)):
                for j in range(i + 1, len(ocr_items) + 1):
                    # 拼接 OCR 内容项并去掉标点符号和空格
                    combined = ''.join([re.sub(r'[^\w]', '', item).lower() for item in ocr_items[i:j]])
                    if combined == target_item_clean:
                        found = True
                        break
                if found:
                    break
        # 如果仍未找到匹配项，记录不一致项
        if not found:
            diff_items.append(target_item)
    # 如果没有不一致项，返回 True
    return len(diff_items) == 0, diff_items

def compare_gift_packages(ocr_file, target_file):
    """
    对比两个表格中的礼包内容，并将结果保存到新的 Excel 文件中。
    """
    # 读取 OCR 结果表格
    ocr_df = pd.read_excel(ocr_file, sheet_name='Sheet1')
    # 读取目标表格
    target_df = pd.read_excel(target_file, sheet_name='Sheet1')

    # 创建一个新的 Excel 文件
    wb = Workbook()
    ws = wb.active
    ws.title = "对比结果"

    # 设置表头
    headers = ["礼包名称", "OCR 内容", "目标内容", "是否一致", "不一致项", "OCR 是否重复", "目标是否重复", "是否重复一致", "OCR 兑换金额", "目标兑换金额", "兑换金额一致"]
    ws.append(headers)

    # 设置颜色填充（红色用于标记不一致的项）
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 遍历 OCR 结果表格中的每一行
    for _, ocr_row in ocr_df.iterrows():
        ocr_gift_name = ocr_row['礼包名称']
        ocr_gift_content = ocr_row['礼包内容']
        ocr_tip_content = ocr_row['tip 内容']

        # 检查是否需要排除该礼包
        exclude = False
        for category, keywords in filter_rules["exclude_keywords"].items():
            if any(keyword in ocr_gift_name for keyword in keywords):
                print(f"跳过礼包: {ocr_gift_name}（包含排除关键词: {category}）")
                exclude = True
                break
        if exclude:
            continue

        # 解析 tip 内容
        ocr_repeat, ocr_amount = parse_tip_content(ocr_tip_content)

        # 在目标表格中查找相同礼包名称的行
        target_row = target_df[target_df['礼包名称'] == ocr_gift_name]
        if not target_row.empty:
            target_gift_content = target_row.iloc[0]['礼包内容（标记红为重点展示道具）']
            target_repeat = target_row.iloc[0]['是否重复']
            target_amount = target_row.iloc[0]['兑换金额']

            # 解析礼包内容
            ocr_items = parse_gift_content(ocr_gift_content)
            target_items = parse_gift_content(target_gift_content)

            # 对比礼包内容
            is_match, diff_items = compare_gift_content(ocr_items, target_items)

            # 对比是否重复和兑换金额
            is_repeat_match = str(ocr_repeat) == str(target_repeat)  # 确保数据类型一致
            if ocr_amount == 0 and target_amount == "免费":
                is_amount_match = True
            elif str(ocr_amount) == str(target_amount):
                is_amount_match = True
            else:
                is_amount_match = False

            # 将结果写入 Excel
            ws.append([
                ocr_gift_name,
                ', '.join(ocr_items),
                ', '.join(target_items),
                is_match,
                ', '.join(diff_items),
                ocr_repeat,
                target_repeat,
                is_repeat_match,
                ocr_amount,
                target_amount,
                is_amount_match
            ])

            # 如果不一致，标记相应的列为红色
            if not is_match:
                ws.cell(row=ws.max_row, column=4).fill = red_fill  # 礼包内容是否一致
            if not is_repeat_match:
                ws.cell(row=ws.max_row, column=8).fill = red_fill  # 是否重复一致
            if not is_amount_match:
                ws.cell(row=ws.max_row, column=11).fill = red_fill  # 兑换金额一致
        else:
            # 如果目标表格中没有对应的礼包名称，标记为不一致
            ws.append([ocr_gift_name, ocr_gift_content, '未找到', False, '全部不一致', ocr_repeat, '未找到', False, ocr_amount, '未找到', False])
            ws.cell(row=ws.max_row, column=4).fill = red_fill  # 礼包内容是否一致
            ws.cell(row=ws.max_row, column=8).fill = red_fill  # 是否重复一致
            ws.cell(row=ws.max_row, column=11).fill = red_fill  # 兑换金额一致

    # 保存 Excel 文件
    comparison_results_file = config["comparison_results_file"]
    wb.save(comparison_results_file)
    print(f"对比结果已保存到 {comparison_results_file}")

if __name__ == "__main__":
    with open('config.json', 'r', encoding='utf-8') as f:
        config = json.load(f)
    # 文件路径
    ocr_output_file = config["ocr_results_file"]   # OCR 结果文件
    target_file = config["target_excel_file"]   # 目标文件
    # 对比礼包内容
    compare_gift_packages(ocr_output_file, target_file)