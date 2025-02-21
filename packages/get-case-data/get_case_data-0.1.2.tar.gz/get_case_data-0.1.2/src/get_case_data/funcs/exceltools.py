import datetime
import os
import openpyxl


def create_excel_file(header, contents, filename):

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for col_num, header_item in enumerate(header, 1):
        sheet.cell(row=1, column=col_num, value=header_item)

    for row_num, row_data in enumerate(contents, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=cell_data)

    workbook.save(datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + f"_{filename}.xlsx")
