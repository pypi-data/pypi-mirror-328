from src.yoklama.yariyil import *
from src.yoklama.ders import *
from src.yoklama.ogrenciler import *
from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.cell import get_column_letter

buyugeDonusum = {ord(u'ı'):u'I', ord(u'i'):u'İ'}
kucugeDonusum = {ord(u'I'):u'ı', ord(u'İ'):u'i'}

thin_border = Border(left=Side(border_style='thin', color='000000'), 
                     right=Side(border_style='thin', color='000000'), 
                     top=Side(border_style='thin', color='000000'), 
                     bottom=Side(border_style='thin', color='000000'))

def kucuktenBuyuge(metin: str) -> str:
    return str.upper(metin.translate(buyugeDonusum))

def buyuktenKucuge(metin: str) -> str:
    return str.lower(metin.translate(kucugeDonusum))

class Yoklama:
    
    def __init__(self, kaynakDosyaYolu, hedefDosyaYolu, universiteAdi, birimAdi, bolumAdi, donem, yariyilBaslangicTarihiMetni, yariyilBitisTarihiMetni, yariyilTatilGunleri, dersAdi, dersKisaAdi, dersSorumlusu, dersKodu, dersinHaftalikProgrami):
        self.kaynakDosyaYolu = kaynakDosyaYolu
        self.hedefDosyaYolu = hedefDosyaYolu
        self.universiteAdi = universiteAdi
        self.birimAdi = birimAdi
        self.bolumAdi = bolumAdi
        self.yariyil = Yariyil(donem, yariyilBaslangicTarihiMetni, yariyilBitisTarihiMetni, yariyilTatilGunleri)
        self.ders = Ders(dersAdi, dersKisaAdi, dersSorumlusu, dersKodu, dersinHaftalikProgrami)
        self.ogrenciler = self.__ogrencileriOku()
        self.calismaKitabi = Workbook()
        self.calismaSayfasi = self.calismaKitabi.active
        self.calismaSayfasi.title = f"{self.ders.kisaAd} - Yoklama"
    
    @property
    def toplamSutun(self):
        return 5 + self.yariyil.toplamHafta * len(self.ders.haftalikProgram)
    
    def __ogrencileriOku(self) -> pd.DataFrame:
        ogrenciler = pd.read_excel(self.kaynakDosyaYolu)
        ogrenciler.columns = ["Öğrenci No", "Ad", "Soyad", "Sınıf", "Alış Tipi", "Not", "Fakülte", "Program"]
        return ogrenciler

    def yoklamaDosyasiOlustur(self):

        def logoEkle():
            logo = Image("Yoklama/agrilogo.png")
            logo.width = 68.4
            logo.height = 54.3
            logo.anchor = "A2"
            self.calismaSayfasi.add_image(logo)
        
        def ogrenciTabloBasliklariniEkle():
            ogrenciTabloBasliklari = ["No", "Öğr.No", "Ad", "Soyad", "S/A"]
            for i in range(len(ogrenciTabloBasliklari)):
                hucre = self.calismaSayfasi.cell(row=1, column=i+1)
                hucre.value = ogrenciTabloBasliklari[i]
                hucre.font = Font(name='Calibri', size=9, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
                self.calismaSayfasi.merge_cells(start_row=1, start_column=i+1, end_row=2, end_column=i+1)
                hucre.alignment = Alignment(horizontal="center", vertical="center")
        
        def haftaTabloBasliklariniEkle():
            for i in range(self.yariyil.toplamHafta):
                hucre = self.calismaSayfasi.cell(row=1, column=6+len(self.ders.haftalikProgram)*i)
                hucre.value = f"{i+1}. Hafta"
                hucre.font = Font(name='Calibri', size=9, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
                self.calismaSayfasi.merge_cells(start_row=1, start_column=6+len(self.ders.haftalikProgram)*i, end_row=1, end_column=6+len(self.ders.haftalikProgram)*(i+1)-1)
                hucre.alignment = Alignment(horizontal="center", vertical="center")
        
        def haftaGunTabloBasliklariniVeTatilGunleriniEkle():
            crow = 2
            ccolumn = 6
            for i in range(self.yariyil.toplamGun):
                gecerliGun = self.yariyil.baslangicTarihi + timedelta(days=i)
                for n in range(len(self.ders.haftalikProgram)):
                    if gecerliGun.weekday() == self.ders.haftalikProgram[n]["haftaGunSirasi"]:
                        hucre = self.calismaSayfasi.cell(row=crow, column=ccolumn)
                        hucre.value = datetime.strftime(gecerliGun, "%d/%m")
                        hucre.font = Font(name='Calibri', size=9, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
                        hucre.alignment = Alignment(horizontal="center", vertical="center",text_rotation=0)
                        dersBaslangic = gecerliGun
                        dersBaslangic += timedelta(hours = self.ders.haftalikProgram[n]["baslangicSaati"].hour)
                        dersBaslangic += timedelta(minutes = self.ders.haftalikProgram[n]["baslangicSaati"].minute)
                        dersBitis = dersBaslangic
                        dersBitis += timedelta(minutes = 60 * self.ders.haftalikProgram[n]["dersSayisi"])
                        for tatilGunu in self.yariyil.tatilGunleri:
                            tatilBaslangic = tatilGunu["baslangicTarihi"]
                            tatilBitis = tatilGunu["baslangicTarihi"] + timedelta(days=tatilGunu["gunSayisi"])
                            if dersBaslangic > tatilBaslangic and dersBitis < tatilBitis:
                                 self.calismaSayfasi.merge_cells(start_row=crow + 1, start_column=ccolumn, end_row=crow + self.ogrenciler.shape[0], end_column=ccolumn)
                                 hucre = self.calismaSayfasi.cell(row=crow + 1, column=ccolumn)
                                 hucre.value = tatilGunu["adi"]
                                 hucre.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
                        ccolumn += 1
        
        def ogrenciBilgileriniEkle():
            self.ogrenciler.reset_index(inplace=True)
            for i in range(self.ogrenciler.shape[0]):
                for j in range(5):
                    hucre = self.calismaSayfasi.cell(row=3+i, column=j+1)
                    if j == 0:
                        hucre.value = self.ogrenciler.iloc[i,j] + 1
                    elif j == 4:
                        hucre.value = str(self.ogrenciler.iloc[i,j]) + "/" + self.ogrenciler.iloc[i,j+1][0]
                    else:
                        hucre.value = self.ogrenciler.iloc[i,j]
                    hucre.font = Font(name='Calibri', size=9, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
                    hucre.alignment = Alignment(horizontal="center", vertical="center")
            

        def hucreGenislikleriniAyarla():
            self.calismaSayfasi.column_dimensions['A'].width = 2.8
            self.calismaSayfasi.column_dimensions['B'].width = 10
            self.calismaSayfasi.column_dimensions['C'].width = 10
            self.calismaSayfasi.column_dimensions['D'].width = 10
            self.calismaSayfasi.column_dimensions['E'].width = 3.5
            for i in range(6,self.toplamSutun+1):
                self.calismaSayfasi.column_dimensions[get_column_letter(i)].width = 12

        def hucreYukseklikleriniAyarla():
            #self.calismaSayfasi.row_dimensions[8].height = 34.5
            pass
        
        def kenarliklariEkle():
            for i in range(self.ogrenciler.shape[0]+2):
                for j in range(self.toplamSutun):
                    hucre = self.calismaSayfasi.cell(row=1+i, column=j+1)
                    hucre.border = thin_border

        def ustBilgiEkle():
            font = 'Calibri,Bold'
            size = 9
            color = '000000'
            baslik1 = ""
            if self.yariyil.donem == 1:
                yil = self.yariyil.baslangicTarihi.year
                baslik1 = f"{yil}-{yil+1} EĞİTİM-ÖĞRETİM YILI\nGÜZ YARIYILI"
            else:
                yil = self.yariyil.bitisTarihi.year
                baslik1 = f"{yil-1}-{yil} EĞİTİM-ÖĞRETİM YILI\nBAHAR YARIYILI"
            self.calismaSayfasi.oddHeader.left.text = baslik1
            self.calismaSayfasi.oddHeader.left.font = font
            self.calismaSayfasi.oddHeader.left.size = size
            self.calismaSayfasi.oddHeader.left.color = color
            self.calismaSayfasi.evenHeader.left.text = baslik1
            self.calismaSayfasi.evenHeader.left.font = font
            self.calismaSayfasi.evenHeader.left.size = size
            self.calismaSayfasi.evenHeader.left.color = color
            baslik2 = f"{kucuktenBuyuge(self.ders.ad)}\n({self.ders.kod})\nDERS YOKLAMA LİSTESİ"
            if len(baslik2) > 63:
                 baslik2 = f"{kucuktenBuyuge(self.ders.kisaAd)}\n({self.ders.kod})\nDERS YOKLAMA LİSTESİ"
            self.calismaSayfasi.oddHeader.center.text = baslik2
            self.calismaSayfasi.oddHeader.center.font = font
            self.calismaSayfasi.oddHeader.center.size = size
            self.calismaSayfasi.oddHeader.center.color = color
            self.calismaSayfasi.evenHeader.center.text = baslik2
            self.calismaSayfasi.evenHeader.center.font = font
            self.calismaSayfasi.evenHeader.center.size = size
            self.calismaSayfasi.evenHeader.center.color = color
            
            baslik3 = f"{kucuktenBuyuge(self.universiteAdi)}\n{kucuktenBuyuge(self.birimAdi)}\n{kucuktenBuyuge(self.bolumAdi)}"
            self.calismaSayfasi.oddHeader.right.text = baslik3
            self.calismaSayfasi.oddHeader.right.font = font
            self.calismaSayfasi.oddHeader.right.size = size
            self.calismaSayfasi.oddHeader.right.color = color
            self.calismaSayfasi.evenHeader.right.text = baslik3
            self.calismaSayfasi.evenHeader.right.font = font
            self.calismaSayfasi.evenHeader.right.size = size
            self.calismaSayfasi.evenHeader.right.color = color

            sayfabilgisi = "&P/&N"
            self.calismaSayfasi.oddFooter.center.text = sayfabilgisi
            self.calismaSayfasi.oddFooter.center.font = font
            self.calismaSayfasi.oddFooter.center.size = size
            self.calismaSayfasi.oddFooter.center.color = color
            self.calismaSayfasi.evenFooter.center.text = sayfabilgisi
            self.calismaSayfasi.evenFooter.center.font = font
            self.calismaSayfasi.evenFooter.center.size = size
            self.calismaSayfasi.evenFooter.center.color = color

            dersSorumlusu = f"Ders Sorumlusu: {self.ders.sorumlu}"
            self.calismaSayfasi.oddFooter.right.text = dersSorumlusu
            self.calismaSayfasi.oddFooter.right.font = font
            self.calismaSayfasi.oddFooter.right.size = size
            self.calismaSayfasi.oddFooter.right.color = color
            self.calismaSayfasi.evenFooter.right.text = dersSorumlusu
            self.calismaSayfasi.evenFooter.right.font = font
            self.calismaSayfasi.evenFooter.right.size = size
            self.calismaSayfasi.evenFooter.right.color = color

        def yazdirmaSecenekleriniAyarla():
            self.calismaSayfasi.print_options.horizontalCentered = True
            self.calismaSayfasi.print_options.verticalCentered = True
            self.calismaSayfasi.print_title_cols = 'A:E'
            self.calismaSayfasi.print_title_rows = '1:2'
            self.calismaSayfasi.print_area = f'A1:{get_column_letter(self.toplamSutun)}{self.ogrenciler.shape[0]+2}'
            self.calismaSayfasi.page_setup.orientation = self.calismaSayfasi.ORIENTATION_LANDSCAPE
            self.calismaSayfasi.page_setup.paperSize = self.calismaSayfasi.PAPERSIZE_A4
            toInch = 0.39370078740157
            self.calismaSayfasi.page_margins = PageMargins(left=0.6 * toInch, right=0.6 * toInch, top=1.8 * toInch, bottom=0.9 * toInch, header=0.5 * toInch, footer=0.5 * toInch)
        
        def calismaKitabiniKaydet():
            self.calismaKitabi.save(self.hedefDosyaYolu)
        
        #logoEkle()
        ogrenciTabloBasliklariniEkle()
        haftaTabloBasliklariniEkle()
        haftaGunTabloBasliklariniVeTatilGunleriniEkle()
        ogrenciBilgileriniEkle()
        hucreGenislikleriniAyarla()
        hucreYukseklikleriniAyarla()
        kenarliklariEkle()
        ustBilgiEkle()
        yazdirmaSecenekleriniAyarla()

        calismaKitabiniKaydet()
        