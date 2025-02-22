from openpyxl import load_workbook
from .format_date import get_format_date_py

def extract_excel_dates(file_path):
    wb = load_workbook(file_path, data_only=False)  
    ws = wb.active  
    checked_columns = set()
    # get headers
    headers = {cell.column: cell.value for cell in ws[1]}  
    result = []
    for row in ws.iter_rows(min_row=2): 
        for cell in row:
            col_index = cell.column  
            if col_index not in checked_columns and cell.is_date:  
                cell_format = cell.number_format  # Format angka
                header_name = headers.get(col_index, f"Kolom {col_index}")  # Ambil header berdasarkan indeks kolom
                result.append({
                    "header": header_name,
                    "format": get_format_date_py(cell_format)
                })
                checked_columns.add(col_index)  
        # stop loop
        if len(checked_columns) == ws.max_column:
            break
    return result


def extract_excel_date_with_pandas(file_path, df):
    extrac_date = extract_excel_dates(file_path)
    for date in extrac_date:
        df[date["header"]] = df[date["header"]].dt.strftime(date["format"])
    return df