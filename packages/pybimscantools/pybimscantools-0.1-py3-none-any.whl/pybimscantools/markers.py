import json
import os

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

from pybimscantools import helper
from pybimscantools import transformations


def apply_transformation_matrix_to_markers_xlsx_and_copy(path: str,
                                                         t: np.array = np.identity(4),
                                                         file_name: str = 'markers.xlsx',
                                                         pcs_description: str = 'IFC') -> None:
    """
    Transform an existing markers.xlsx file with a transformation matrix T and save it as markers_ifc.xlsx.
    It is assumed that the markers.xlsx file exists in the directory path/markers.
    Further it is assumed that T transforms the marker points from an arbitrary frame to ifc.
    """

    path = os.path.join(path, "markers")

    file_was_found = helper.does_file_name_exist_in_path(path, file_name)
    if not file_was_found:
        print(f"   no {file_name} file was found")
        return

    file = os.path.join(path, file_name)
    file_ifc = os.path.join(path, os.path.splitext(file_name)[0] + '_ifc' + os.path.splitext(file_name)[1])
    if os.name == 'nt':  # windows
        cmd = f'copy "{file}" "{file_ifc}"'
    else:  # unix/linux
        cmd = f'cp "{file}" "{file_ifc}"'
    os.system(cmd)

    workbook = openpyxl.load_workbook(file_ifc)
    sheet = workbook.active

    sheet['C2'] = pcs_description

    last_row = sheet.max_row
    markers_left = np.array([[sheet.cell(row=i, column=j).value for j in range(5, 8)] for i in range(4, last_row + 1)])
    markers_right = np.array([[sheet.cell(row=i, column=j).value for j in range(2, 5)] for i in range(4, last_row + 1)])

    for i in range(len(markers_left)):
        markers_left[i] = t[0:3, 0:3].dot(markers_left[i]) + t[0:3, 3]
        markers_right[i] = t[0:3, 0:3].dot(markers_right[i]) + t[0:3, 3]

    for i in range(markers_right.shape[0]):
        for j in range(markers_right.shape[1]):
            sheet.cell(row=i + 4, column=j + 2).value = markers_right[i, j]
            sheet.cell(row=i + 4, column=j + 5).value = markers_left[i, j]

    workbook.save(file_ifc)


def convert_markers_from_xlsx_to_json(path: str,
                                      file_name: str = 'markers_ifc.xlsx') -> None:
    """
    Takes an existing markers_ifc.xlsx file and converts it to marker_nr_*.json files and stores these in the
    directory path/markers/json.
    It is assumed that the markers_ifc.xlsx file exists in the directory path/markers.
    """

    path = os.path.join(path, "markers")

    path_sub_dir = helper.create_subdir_if_not_exists(path, "json")

    file_was_found = helper.does_file_name_exist_in_path(path, file_name)
    if not file_was_found:
        print(f"   no {file_name} file was found")
        return

    info_table = pd.read_excel(os.path.join(path, file_name), usecols='A:G', nrows=2, engine='openpyxl')

    version = info_table[info_table.columns[0]][0]
    pcs_description = info_table[info_table.columns[2]][0]
    is_valid = False
    if info_table[info_table.columns[1]][0] == 'true':
        is_valid = True
    project_name = info_table[info_table.columns[3]][0]
    project_information = info_table[info_table.columns[4]][0]
    measured_date = info_table[info_table.columns[5]][0]
    contact_person = info_table[info_table.columns[6]][0]


    marker_table = pd.read_excel(os.path.join(path, file_name), usecols='A:G', skiprows=2, engine='openpyxl')

    marker_table['l_xyz'] = marker_table[['l_x', 'l_y', 'l_z']].values.tolist()
    marker_table['r_xyz'] = marker_table[['r_x', 'r_y', 'r_z']].values.tolist()

    for index, row in marker_table.iterrows():
        data = {
            "project_name": str(project_name),
            "project_information": str(project_information),
            "version": str(version),
            "marker_nr": int(marker_table['marker_nr'][index]),
            "valid": is_valid,
            "pcs_description": str(pcs_description),
            "l_xyz": marker_table['l_xyz'][index],
            "r_xyz": marker_table['r_xyz'][index],
            "measured_date": str(measured_date),
            "contact_person": str(contact_person)
        }

        marker_name = 'marker_nr_' + str(marker_table['marker_nr'][index]) + '.json'
        with open(os.path.join(path_sub_dir, marker_name), 'w', encoding="utf-8") as json_file:
            json.dump(data, json_file, indent=4)


def read_markers_from_json_to_table(path: str) -> pd.DataFrame:
    """
    Takes a list of marker_nr_*.json files and converts it to a pandas DataFrame.
    It is assumed that the marker_nr_*.json files exist in the directory path/markers/json.
    """

    path = os.path.join(path, "markers/json")

    index = 0
    marker_table = pd.DataFrame({'project_name': pd.Series(dtype='str'),
                                 'project_information': pd.Series(dtype='str'),
                                 'version': pd.Series(dtype='str'),
                                 'marker_nr': pd.Series(dtype='str'),
                                 'valid': pd.Series(dtype='bool'),
                                 'pcs_description': pd.Series(dtype='str'),
                                 'l_x': pd.Series(dtype='float'),
                                 'l_y': pd.Series(dtype='float'),
                                 'l_z': pd.Series(dtype='float'),
                                 'r_x': pd.Series(dtype='float'),
                                 'r_y': pd.Series(dtype='float'),
                                 'r_z': pd.Series(dtype='float'),
                                 'measured_date': pd.Series(dtype='str'),
                                 'contact_person': pd.Series(dtype='str')})
    for file in os.listdir(path):
        if file[0:10] == "marker_nr_":  # marker_nr_
            marker_i_table = pd.read_json(os.path.join(path, file), orient='records')

            # write all the data to one single row
            marker_table.at[index, 'project_name'] = marker_i_table.at[0, 'project_name']
            marker_table.at[index, 'project_information'] = marker_i_table.at[0, 'project_information']
            marker_table.at[index, 'version'] = marker_i_table.at[0, 'version']
            marker_table.at[index, 'marker_nr'] = marker_i_table.at[0, 'marker_nr']
            marker_table.at[index, 'valid'] = marker_i_table.at[0, 'valid']
            marker_table.at[index, 'pcs_description'] = marker_i_table.at[0, 'pcs_description']
            marker_table.at[index, 'l_x'] = marker_i_table.at[0, 'l_xyz']
            marker_table.at[index, 'l_y'] = marker_i_table.at[1, 'l_xyz']
            marker_table.at[index, 'l_z'] = marker_i_table.at[2, 'l_xyz']
            marker_table.at[index, 'r_x'] = marker_i_table.at[0, 'r_xyz']
            marker_table.at[index, 'r_y'] = marker_i_table.at[1, 'r_xyz']
            marker_table.at[index, 'r_z'] = marker_i_table.at[2, 'r_xyz']
            marker_table.at[index, 'measured_date'] = marker_i_table.at[0, 'measured_date']
            marker_table.at[index, 'contact_person'] = marker_i_table.at[0, 'contact_person']
            index = index + 1

    if index == 0:
        print("   no marker_nr_*.json files were found")
        return marker_table

    marker_table = marker_table.sort_values(by=['marker_nr'])

    return marker_table


def apply_transformation_matrix_to_marker_table(marker_table: pd.DataFrame,
                                                t: list,
                                                pcs_description: str = 'IFC') -> pd.DataFrame:
    """
    Transform an existing marker table with a transformation matrix T and return the transformed marker table.
    """

    markers_left = marker_table[['l_x', 'l_y', 'l_z']].to_numpy()
    markers_right = marker_table[['r_x', 'r_y', 'r_z']].to_numpy()

    for i in range(len(markers_left)):
        markers_left[i] = t[0:3, 0:3].dot(markers_left[i]) + t[0:3, 3]
        markers_right[i] = t[0:3, 0:3].dot(markers_right[i]) + t[0:3, 3]

    marker_table[['l_x', 'l_y', 'l_z']] = markers_left
    marker_table[['r_x', 'r_y', 'r_z']] = markers_right

    marker_table['pcs_description'] = pcs_description

    return marker_table


def convert_relative_corners_tag_from_xlsx_to_json(path: str,
                                                   file_name: str = 'relative_corners_tag.xlsx') -> str:
    """
    Takes an existing relative_corners_tag.xlsx file and converts it to relative_corners_tag.json and stores it in the
    directory path/markers/json.
    """

    path = os.path.join(path, "markers")

    path_sub_dir = helper.create_subdir_if_not_exists(path, "json")

    file_was_found = helper.does_file_name_exist_in_path(path, file_name)
    if not file_was_found:
        print(f"   no {file_name} file was found")
        return

    corner_table = pd.read_excel(os.path.join(path, file_name), usecols='B:F', engine='openpyxl')

    file_name_json = os.path.splitext(file_name)[0] + '.json'
    corner_table.to_json(os.path.join(path_sub_dir, file_name_json), orient='records', indent=4)

    return file_name_json


def read_relative_corners_tag_from_json_to_table(path: str,
                                                 file_name: str = 'relative_corners_tag.json') -> pd.DataFrame:
    """
    Takes a relative_corners_tag.json file and converts it to a pandas DataFrame.
    """

    path = os.path.join(path, "markers/json")

    return pd.read_json(os.path.join(path, file_name), orient='records')


def get_chilli_tag_corners_from_table_as_list(chilli_tag_table: pd.DataFrame) -> list:
    """
    Takes a tag_corners.json file and converts it to a list of corner points.
    """

    corner_point_list = []
    for marker_nr in np.unique(np.sort(chilli_tag_table["marker_nr"])):
        sub_table = chilli_tag_table[chilli_tag_table["marker_nr"] == marker_nr]
        x = sub_table["x"].values
        y = sub_table["y"].values
        z = sub_table["z"].values
        for i in range(len(x)):
            corner_point_list.append([x[i], y[i], z[i]])

    return corner_point_list


def create_tag_corners_json(path: str,
                            marker_table: pd.DataFrame,
                            corner_table: pd.DataFrame,
                            file_name: str = 'corners_tag.json') -> pd.DataFrame:
    """
    Takes a marker table and a corner table and converts it to a tag_corners.json file and stores it in the
    directory path/markers/json.
    """

    path = os.path.join(path, "markers/json")

    markers_right = marker_table[['r_x', 'r_y', 'r_z']].to_numpy()
    markers_left = marker_table[['l_x', 'l_y', 'l_z']].to_numpy()
    marker_nr = marker_table['marker_nr'].to_numpy()
    relative_corners = corner_table[['rel_cc_x', 'rel_cc_y', 'rel_cc_z']].to_numpy()

    tags = np.zeros((markers_right.shape[0] * 4, 5))

    for j in range(markers_right.shape[0]):

        # calculate rotation matrix
        axis_x = (markers_left[j, :] - markers_right[j, :])
        axis_y = np.array([-axis_x[1], axis_x[0], 0])
        axis_x = axis_x / np.linalg.norm(axis_x)
        axis_y = axis_y / np.linalg.norm(axis_y)
        axis_z = np.cross(axis_x, axis_y)
        r = np.vstack([axis_x, axis_y, axis_z]).transpose()

        for i in range(relative_corners.shape[0]):
            tags[(j * 4) + i, 0] = marker_nr[j]
            tags[(j * 4) + i, 1] = i
            tags[(j * 4) + i, 2:5] = (r @ relative_corners[i, :].transpose() + markers_right[j, :]).transpose()

    tag_table = pd.DataFrame(tags, columns=['marker_nr', 'corner_nr', 'x', 'y', 'z'])

    tag_table['marker_nr'] = tag_table['marker_nr'].astype(int)
    tag_table['corner_nr'] = tag_table['corner_nr'].astype(int)

    tag_table.to_json(os.path.join(path, file_name), orient='records', indent=4)

    return tag_table


def create_tag_corners_txt_from_table(path: str,
                                      tag_table: pd.DataFrame,
                                      file_name: str = 'px4tagger.txt') -> None:
    """
    Takes a tag_corners.json file and converts it to a px4tagger.txt file and stores it in the
    directory path/markers/json.
    """

    tag_table.to_csv(os.path.join(path, file_name), index=False, header=False)


def read_marker_corners_json_to_table(path: str,
                                      file_name: str = 'tag_corners.json') -> pd.DataFrame:
    """
    Takes a tag_corners.json file and converts it to a pandas DataFrame.
    """

    path = os.path.join(path, "markers/json")

    return pd.read_json(os.path.join(path, file_name), orient='records')


def plot_tags(tag_table: pd.DataFrame, color: str = 'b', marker: str = 'o') -> None:
    """
    Plot the coordinates of the tags_coords list in 3D assuming
    it is a list of tags consisting of sets of 4 coordinates
    """
    tag_coords = get_chilli_tag_corners_from_table_as_list(tag_table)

    # plot the coordinates in 3D
    fig = plt.figure()
    ax = fig.add_subplot(projection='3d')

    ax.set_box_aspect([1.0, 1.0, 1.0])

    # create a list that contains the mean of group of always 4 coordinates
    for i in range(0, len(tag_coords), 4):
        tag_coordinates = np.array(tag_coords[i:i + 4])
        tag_center = np.mean(tag_coordinates, axis=0)

        # retrieve the marker number from the tag_table
        marker_nr = tag_table.iloc[i]['marker_nr']

        axis_y = transformations.get_normal_to_3d_points(tag_coordinates)
        axis_x = tag_coordinates[0, :] - tag_coordinates[1, :]
        axis_x = axis_x / np.linalg.norm(axis_x)
        axis_z = np.cross(axis_x, axis_y)
        if axis_z[2] < 0:
            axis_y = -axis_y
            axis_z = -axis_z

        # draw tag corner 0 blue
        ax.scatter(tag_coordinates[0,0], tag_coordinates[0,1], tag_coordinates[0,2],
                    c='b', marker=marker)
        # draw the other 3 corners black
        ax.scatter(tag_coordinates[1:, 0], tag_coordinates[1:, 1], tag_coordinates[1:, 2],
                    c='k', marker=marker)

        # draw a frame with the tag convention (x- left, y- out, z- up)
        ax.quiver(tag_center[0], tag_center[1], tag_center[2],
                    axis_x[0], axis_x[1], axis_x[2], length=1.0, color='r')
        ax.quiver(tag_center[0], tag_center[1], tag_center[2],
                    axis_y[0], axis_y[1], axis_y[2], length=1.0, color='g')
        ax.quiver(tag_center[0], tag_center[1], tag_center[2],
                    axis_z[0], axis_z[1], axis_z[2], length=1.0, color='b')
        
        # add marker number text at the tag center
        ax.text(tag_center[0], tag_center[1], tag_center[2], f'{int(marker_nr)}', color='black')

    x_limits = ax.get_xlim3d()
    y_limits = ax.get_ylim3d()
    z_limits = ax.get_zlim3d()

    x_range = abs(x_limits[1] - x_limits[0])
    x_middle = np.mean(x_limits)
    y_range = abs(y_limits[1] - y_limits[0])
    y_middle = np.mean(y_limits)
    z_range = abs(z_limits[1] - z_limits[0])
    z_middle = np.mean(z_limits)

    # the plot bounding box is a sphere in the sense of the infinity
    # norm, hence I call half the max range the plot radius.
    plot_radius = 0.5 * max([x_range, y_range, z_range])

    ax.set_xlim3d([x_middle - plot_radius, x_middle + plot_radius])
    ax.set_ylim3d([y_middle - plot_radius, y_middle + plot_radius])
    ax.set_zlim3d([z_middle - plot_radius, z_middle + plot_radius])

    # set the labels
    ax.set_xlabel('x')
    ax.set_ylabel('y')
    ax.set_zlabel('z')

    plt.show()