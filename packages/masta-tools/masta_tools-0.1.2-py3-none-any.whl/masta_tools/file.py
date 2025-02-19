# -*- coding: utf-8 -*-
import csv, os
from itertools import count
import xlrd
from openpyxl import load_workbook, Workbook

class File:
    def __init__(self, filepath):
        self.filepath = filepath
        self.filename = os.path.basename(filepath)
        self.data = None
        
        
    @classmethod
    def new(cls, data, filename, directory=None, overwrite=True, includeHdr=True):
        """
        directory will default to downloads folder if left blank
        """
        if directory is None:
            directory = cls.download_dir()
            pass
        filepath = os.path.join(directory, filename)
        if os.path.exists(filepath) and not overwrite:
            raise Exception('This path exists already. Set overwrite true if you want to overwrite')
        obj=cls(filepath)
        obj.data=data
        obj.export_csv(filepath, data, includeHdr=includeHdr)
        return obj
        
    
    @staticmethod
    def download_dir():
        return os.path.join(os.getenv('USERPROFILE'), 'Downloads')

    
    @staticmethod
    def remap(fieldnames):
        uq={col:count(1) for col in list(set(fieldnames))}
        done=[]
        new=[]
        for col in fieldnames:
            ct=""
            if col in done:
                ct="-{num}".format(num=next(uq[col]))
            else:
                done.append(col)
            new.append('{column}{count}'.format(column=col, count=ct))
        return new
    
    
    

    ### CSV
    @staticmethod
    def load_csv(filepath, encoding='utf-8', delim=","):
        with open(filepath, encoding=encoding) as f:
            reader=csv.reader(f, delimiter=delim)
            fieldnames=File.remap(next(reader))
            data=[dict(zip(fieldnames,row)) for row in reader]
        return data
    
    @staticmethod
    def as_text(filepath, encoding='utf-8'):
        with open(filepath, 'r', encoding=encoding) as f:
            text = f.read()
        return text
    
    @staticmethod
    def load_tsv(filepath, encoding='utf-8', colmask=0):
        with open(filepath, encoding=encoding) as f:
            text = f.read()
            data = [[i for i in x.split('\t')] for x in text.split('\n')]
            hdrs=data[0]
            if colmask>0:
                hdrs=hdrs[:colmask]
            hdrs=File.remap(hdrs)
            func=lambda x: True if colmask==0 else len(x)>=colmask
            data = [dict(zip(hdrs,x[:colmask]))for x in data[1:] if func(x)]
        return data
    
    
    @staticmethod
    def export_csv(filepath, data, includeHdr=True):
        with open(filepath, 'w', newline='') as f:
            writer=csv.DictWriter(f, list(data[0].keys()))
            if includeHdr:
                writer.writeheader()
            writer.writerows(data)
         
            
    @staticmethod
    def read_xl(filepath):
        if str(filepath).lower().endswith('.xls'):
            wb=xlrd.open_workbook(filepath)
            sheets = wb.sheets()
            output = {}
            for sheet in sheets:
                output[sheet.name]=[sheet.row_values(x) for x in range(sheet.nrows)]
        else:
            wb=load_workbook(filepath, read_only=True,
                             data_only=True, keep_links=False)
            output = {}
            for sheet in wb.worksheets:
                output[sheet.title]=[x for x in sheet.values]
        return output
    
    @staticmethod
    def create_xls(data, title, filepath):
        wb=Workbook()
        ws = wb.active
        ws.title = title
        fields = list(data[0].keys())
        ws.append(fields)
        for row in data:
            ws.append(list(row.values()))
        wb.save(filepath)
        wb.close()
        file = File(filepath)
        file.data = data
        return file


    ### GENERAL
    @staticmethod
    def delete(filepath):
        try:
            os.remove(filepath)
        except:
            print('File does not exist ({})'.format(filepath))    
            
            
    @staticmethod
    def listdir(folderpath):
        files=os.listdir(folderpath)
        return [File(os.path.join(folderpath,x)) for x in files]
    
    
    def __repr__(self):
        return self.filepath